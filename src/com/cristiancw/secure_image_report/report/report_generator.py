import os
from datetime import datetime

import pandas
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

from com.cristiancw.secure_image_report.report.report_content import ReportContent


class ReportGenerator:
    """
    Based in the ReportContent object with ReportContentLine will create a report.
    """

    def __init__(self, content: ReportContent = None) -> None:
        """
        Main constructor.
        :param content: with all the lines to create the report
        """
        cwd = os.getcwd()
        new_directory = 'reports'
        path = os.path.join(cwd, new_directory)
        os.makedirs(path, exist_ok=True)
        self._xls_file = f"reports/secure-image-report-{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx"
        self._content = content

    def generate(self) -> None:
        """
        Generate the report xlsx in the reports/ folder.
        :return: None
        """
        df = pandas.DataFrame(self._content.get_values())
        df.to_excel(self._xls_file, sheet_name='Scan Report Items', index=False, engine='openpyxl')

    def format(self) -> None:
        """
        After creating the raw report apply some format to get more presentable content.
        :return:  None
        """
        workbook = load_workbook(self._xls_file)
        worksheet = workbook.active

        # apply the auto-filter in the report
        worksheet.auto_filter.ref = worksheet.dimensions

        for col, width in self._content.get_column_sizes().items():
            worksheet.column_dimensions[col].width = width
            if col == 'I':
                worksheet.column_dimensions[col].alignment = Alignment(wrap_text=True)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value in self._content.get_cel_style():
                    style = self._content.get_cel_style()[cell.value]
                    cell.fill = style.get('fill', PatternFill())
                    cell.font = style.get('font', Font())
        workbook.save(self._xls_file)
